import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import VideoRecord


def _sheet_name(name: str, used: set) -> str:
    base = re.sub(r'[\\/*?:\[\]]', "_", name)[:28] or "未知作者"
    candidate, index = base, 2
    while candidate in used:
        suffix = f"_{index}"
        candidate = base[: 31 - len(suffix)] + suffix
        index += 1
    used.add(candidate)
    return candidate


def _sort_key(item: VideoRecord) -> datetime:
    for fmt in ("%Y-%m-%d", "%Y%m%d"):
        try:
            return datetime.strptime(item.published_at[:10], fmt)
        except ValueError:
            pass
    return datetime.min


def _file_stem(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|\x00-\x1f]', "_", name or "未知作者").strip(". ")[:80] or "未知作者"


def export_excel(records: Iterable[VideoRecord], target_dir: Path) -> List[Path]:
    grouped = defaultdict(list)
    for record in records:
        if record.status == "done":
            grouped[record.author].append(record)

    target_dir.mkdir(parents=True, exist_ok=True)
    results = []
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["发布时间", "视频标题", "字幕", "视频链接"]
    for author in sorted(grouped):
        book = Workbook()
        sheet = book.active
        sheet.title = _sheet_name(author, set())
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for record in sorted(grouped[author], key=_sort_key, reverse=True):
            sheet.append([record.published_at, record.title, record.transcript, record.video_url])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = [14, 42, 90, 48]
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row_index in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row_index].height = 60
        path = target_dir / f"{_file_stem(author)}_字幕汇总.xlsx"
        book.save(path)
        results.append(path)

    if not grouped:
        book = Workbook()
        sheet = book.active
        sheet = book.create_sheet("结果")
        book.remove(book["Sheet"])
        sheet.append(["提示"])
        sheet.append(["本次没有成功转写的视频，请查看运行报告.json。"])
        path = target_dir / "处理结果.xlsx"
        book.save(path)
        results.append(path)
    return results

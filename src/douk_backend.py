"""Independent DouK-Downloader adapter for author-profile collection."""
import json
import re
import sqlite3
import shutil
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional

from openpyxl import load_workbook

from .models import RunOptions, VideoRecord
from .session_login import session_cookie_header


Log = Callable[[str], None]
PROJECT_DIR = Path(__file__).resolve().parents[1]
VENDOR_DIR = PROJECT_DIR / "vendor" / "TikTokDownloader"
VOLUME_DIR = VENDOR_DIR / "Volume"
VENDOR_PYTHON = VENDOR_DIR / ".venv" / "bin" / "python"
LOCAL_COOKIE = PROJECT_DIR / ".private" / "douyin_cookie.txt"
SENSITIVE_COOKIE = re.compile(
    r"(?i)(sessionid(?:_ss)?|msToken|ttwid|odin_tt|passport_[\\w]+)=([^;\\s]+)"
)


def _safe_log(message: str) -> str:
    if message.lower().startswith("cookie:"):
        return "Cookie：已隐藏"
    return SENSITIVE_COOKIE.sub(r"\\1=已隐藏", message)


def setup_message() -> str:
    return "尚未安装 DouK-Downloader 运行环境。请先双击“安装全部组件.command”。"


def _settings(urls: Iterable[str], root: Path) -> Dict[str, object]:
    cookie = session_cookie_header()
    if not cookie and LOCAL_COOKIE.exists():
        cookie = LOCAL_COOKIE.read_text(encoding="utf-8").strip()
    existing: Dict[str, object] = {}
    settings_path = VOLUME_DIR / "settings.json"
    if settings_path.exists():
        try:
            existing = json.loads(settings_path.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            existing = {}
    existing.update({
        "accounts_urls": [
            {"mark": f"作者_{index}", "url": url, "tab": "", "earliest": "", "latest": "", "enable": True}
            for index, url in enumerate(urls, 1)
        ],
        "root": str(root),
        "folder_name": "Download",
        "name_format": "create_time type nickname desc",
        "storage_format": "xlsx",
        "cookie": cookie,
        "download": True,
        "max_pages": 0,
        "ffmpeg": shutil.which("ffmpeg") or "",
        # 菜单 5：终端交互；子菜单 1：抖音账号作品；再选 1：accounts_urls。
        "run_command": "5 1 1 Q",
    })
    return existing


def _run_downloader(urls: List[str], output_dir: Path, log: Log) -> Path:
    if not VENDOR_PYTHON.exists():
        raise RuntimeError(setup_message())
    destination = output_dir / "douk_downloads"
    destination.mkdir(parents=True, exist_ok=True)
    VOLUME_DIR.mkdir(parents=True, exist_ok=True)
    # 上游首次启动会要求确认免责声明和选择语言，自动任务不能停在这里。
    database_path = VOLUME_DIR / "DouK-Downloader.db"
    with sqlite3.connect(database_path) as database:
        database.execute(
            "CREATE TABLE IF NOT EXISTS config_data (NAME TEXT PRIMARY KEY, VALUE INTEGER NOT NULL)"
        )
        database.execute(
            "CREATE TABLE IF NOT EXISTS option_data (NAME TEXT PRIMARY KEY, VALUE TEXT NOT NULL)"
        )
        database.execute(
            "INSERT OR REPLACE INTO config_data (NAME, VALUE) VALUES ('Disclaimer', 1)"
        )
        database.execute(
            "INSERT OR REPLACE INTO option_data (NAME, VALUE) VALUES ('Language', 'zh_CN')"
        )
    settings_path = VOLUME_DIR / "settings.json"
    settings_path.write_text(
        json.dumps(_settings(urls, destination), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    settings_path.chmod(0o600)
    log("正在调用 DouK-Downloader 采集作者主页作品…")
    process = subprocess.Popen(
        [str(VENDOR_PYTHON), "main.py"], cwd=str(VENDOR_DIR),
        stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True,
        encoding="utf-8", errors="replace",
    )
    assert process.stdout is not None
    for line in process.stdout:
        message = line.strip()
        if message:
            log(f"DouK-Downloader：{_safe_log(message[:500])}")
    if process.wait() != 0:
        raise RuntimeError("DouK-Downloader 执行失败，请查看上方日志及其参数配置。")
    return destination


def _cell(row: tuple, headers: Dict[str, int], name: str) -> str:
    index = headers.get(name)
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value or "").strip()


def _find_media(root: Path, video_id: str, published_at: str) -> Optional[Path]:
    for extension in ("*.mp4", "*.mov", "*.mkv", "*.webm"):
        for file in root.rglob(extension):
            if video_id and video_id in file.name:
                return file
    # DouK-Downloader 默认以“发布时间-类型-作者-标题”命名，大多数文件名不含作品 ID。
    prefix = published_at.replace(":", ".")
    if prefix:
        for extension in ("*.mp4", "*.mov", "*.mkv", "*.webm"):
            for file in root.rglob(extension):
                if file.name.startswith(prefix):
                    return file
    return None


def _records_from_xlsx(root: Path, log: Log) -> List[VideoRecord]:
    records = []
    for workbook_path in root.rglob("*.xlsx"):
        book = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            sheet = book.active
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                continue
            headers = {str(value).strip(): index for index, value in enumerate(header_row) if value}
            if "作品ID" not in headers or "发布时间" not in headers:
                continue
            for row in rows:
                video_id = _cell(row, headers, "作品ID")
                published_full = _cell(row, headers, "发布时间")
                media = _find_media(root, video_id, published_full)
                if not video_id or not media:
                    continue
                records.append(VideoRecord(
                    author=_cell(row, headers, "账号昵称") or media.parent.name,
                    author_url="", video_id=video_id,
                    video_url=_cell(row, headers, "作品链接"),
                    title=_cell(row, headers, "作品描述") or video_id,
                    published_at=published_full[:10],
                    media_path=str(media),
                ))
        finally:
            book.close()
    log(f"已从下载记录匹配到 {len(records)} 个视频文件。")
    return records


def collect_profile_videos(urls: List[str], options: RunOptions, log: Log) -> List[VideoRecord]:
    records = _records_from_xlsx(_run_downloader(urls, options.output_dir, log), log)
    if not records:
        raise RuntimeError("下载器未生成可匹配的视频与 XLSX 记录，请检查 Cookie、主页可访问性和参数配置。")
    return records


def extract_audio(record: VideoRecord, options: RunOptions, log: Log) -> Path:
    source = Path(record.media_path)
    if not source.exists():
        raise RuntimeError("找不到下载的视频文件。")
    target_dir = options.output_dir / "audio" / record.author
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / f"{record.video_id}.mp3"
    if target.exists():
        return target
    log(f"提取音频：{record.title}")
    result = subprocess.run(
        ["ffmpeg", "-y", "-i", str(source), "-vn", "-codec:a", "libmp3lame", "-q:a", "2", str(target)],
        stdout=subprocess.DEVNULL, stderr=subprocess.PIPE, text=True, encoding="utf-8", errors="replace",
    )
    if result.returncode != 0 or not target.exists():
        raise RuntimeError(f"ffmpeg 音频提取失败：{result.stderr[-300:]}")
    return target
